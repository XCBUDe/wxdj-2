#!/usr/bin/env python3
"""
按名单批量采集：读取 Excel 中的 uid 列，对每家经销商实时抓取
店头名称 / 车型报价 / 软文日期 / 头图截图，输出同格式 xlsx。
"""
import sys
import time
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl

# ── 省/城市推断表（从简称关键字推断）────────────────────────────────────
_PROV_CITY = [
    # (省份, 城市关键字列表, 城市名)
    ("福建", ["福清"],          "福州"),
    ("福建", ["福建", "厦门"],   "厦门"),
    ("福建", ["龙岩"],          "龙岩"),
    ("福建", ["泉州", "晋江"],   "泉州"),
    ("福建", ["三明"],          "三明"),
    ("福建", ["漳州"],          "漳州"),
    ("广东", ["东莞"],          "东莞"),
    ("广东", ["佛山"],          "佛山"),
    ("广东", ["广州", "广东"],   "广州"),
    ("广东", ["惠州"],          "惠州"),
    ("广东", ["江门"],          "江门"),
    ("广东", ["揭阳"],          "揭阳"),
    ("广东", ["茂名"],          "茂名"),
    ("广东", ["梅州"],          "梅州"),
    ("广东", ["清远"],          "清远"),
    ("广东", ["汕头"],          "汕头"),
    ("广东", ["韶关"],          "韶关"),
    ("广东", ["深圳"],          "深圳"),
    ("广东", ["湛江"],          "湛江"),
    ("广东", ["肇庆"],          "肇庆"),
    ("广东", ["中山"],          "中山"),
    ("广东", ["珠海"],          "珠海"),
    ("海南", ["海南"],          "海口"),
]


def _infer_prov_city(short_name: str):
    for prov, keywords, city in _PROV_CITY:
        if any(k in short_name for k in keywords):
            return prov, city
    return "未知", "未知"


def load_dealers(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Sheet1"]
    dealers = []
    for r in range(4, ws.max_row + 1):
        uid   = ws.cell(r, 3).value
        short = ws.cell(r, 2).value or ""
        if uid is None or str(uid) == "#N/A" or not str(uid).isdigit():
            continue
        prov, city = _infer_prov_city(short)
        dealers.append({
            "province":  prov,
            "city":      city,
            "dealer_id": str(uid),
            "name":      short,   # 采集后会被真实名称覆盖
        })
    return dealers


def main():
    XLSX_INPUT  = "/root/.claude/uploads/97740e16-e0aa-4c55-a8e7-e869472d0cfa/5b9dab83-H______________.xlsx"
    OUTPUT      = "output/result_full.xlsx"
    CHECKPOINT  = "output/checkpoint_full.jsonl"
    CONCURRENCY = 2

    from src.detail import fetch_dealer_detail
    from src.crawler import _load_checkpoint, _append_checkpoint, _load_all_from_checkpoint
    from src.export_xlsx import export_to_xlsx

    features = {"pricing": True, "screenshot": True, "article_date": True}

    dealers = load_dealers(XLSX_INPUT)
    print(f"名单共 {len(dealers)} 家经销商")

    Path(OUTPUT).parent.mkdir(parents=True, exist_ok=True)

    done_ids = _load_checkpoint(CHECKPOINT)
    todo = [d for d in dealers if d["dealer_id"] not in done_ids]
    print(f"待采集: {len(todo)} 家  已完成: {len(done_ids)} 家")

    records_done = _load_all_from_checkpoint(CHECKPOINT)

    def _work(dealer: dict) -> dict:
        result = fetch_dealer_detail(dealer, features, browser=None)
        _append_checkpoint(CHECKPOINT, result)
        time.sleep(1.5)
        return result

    new_records = []
    with ThreadPoolExecutor(max_workers=CONCURRENCY) as pool:
        futures = {pool.submit(_work, d): d for d in todo}
        for i, fut in enumerate(as_completed(futures), 1):
            d = futures[fut]
            try:
                rec = fut.result()
                new_records.append(rec)
                n = len(rec.get("pricing") or [])
                status = f"OK {n}款" if not rec.get("error") else f"ERR: {rec['error'][:50]}"
                print(f"  [{i:02d}/{len(todo)}] {d['dealer_id']} {rec.get('name') or d['name']} → {status}")
            except Exception as e:
                print(f"  [{i:02d}/{len(todo)}] {d['dealer_id']} FATAL: {e}")

    all_records = records_done + new_records
    export_to_xlsx(all_records, OUTPUT, features)
    print(f"\n完成！共 {len(all_records)} 家，输出: {OUTPUT}")


if __name__ == "__main__":
    main()
