#!/usr/bin/env python3
"""
将 result_full.xlsx 的裸车价与「网销平台报价标准 2026.6.2」对比。

匹配规则（按可靠性排序）：
  1. 车系映射：我方车系 → 标准车系候选（SERIES_MAP）
  2. 在候选内用「指导价 == 标准市场价格」定位车型（厂方指导价稳定、唯一性强）
  3. 同一指导价命中多行时，用车型名称字符重合度消歧

着色：
  - 裸车价 ≠ 标准裸车价(市场价-现金优惠，0容错) → 裸车价单元格标红
  - 车系/车型无法在标准中匹配          → 车系单元格标黄
输出：output/result_full_checked.xlsx
"""
import re
import collections
import openpyxl
from openpyxl.styles import PatternFill

STD_PATH = '/root/.claude/uploads/901fc9bd-ff28-47f1-952b-234cb1c9d3dd/a9490c25-___________2026.6.2_.xlsx'
SRC_PATH = 'output/result_full.xlsx'
OUT_PATH = 'output/result_full_checked.xlsx'

RED = PatternFill('solid', fgColor='FFC7CE')      # 裸车价错误
YELLOW = PatternFill('solid', fgColor='FFEB9C')   # 车系/车型无法匹配标准

# ── 我方车系 → 标准车系候选 ────────────────────────────────────────────
SERIES_MAP = {
    '艾瑞泽5':       ['艾瑞泽5 卓越版'],
    '艾瑞泽8':       ['2025款 艾瑞泽8', '2025款艾瑞泽8卓越版'],
    '艾瑞泽8 PRO':   ['艾瑞泽8 PRO', '艾瑞泽8 PRO 400T'],
    '瑞虎3x':        ['瑞虎3x 卓越版'],
    '瑞虎5':         ['瑞虎5'],
    '瑞虎5x':        ['瑞虎5x高能版', '瑞虎5x 卓越版'],
    '瑞虎7':         ['瑞虎7卓越版\n（全新一代瑞虎7）', '瑞虎7 高能版', '全新瑞虎7', '全新瑞虎7 C-DM'],
    '瑞虎7 PLUS':    ['瑞虎7 PLUS'],
    '瑞虎7L':        ['瑞虎7L'],
    '瑞虎8':         ['第五代瑞虎8', '瑞虎8 卓越版'],
    '瑞虎8 PLUS':    ['全新瑞虎8 PLUS'],
    '瑞虎8 PLUS C-DM': ['全新瑞虎8 PLUS C-DM'],
    '瑞虎8 PRO':     ['瑞虎8 PRO'],
    '瑞虎8L':        ['瑞虎8 L'],
    '瑞虎9':         ['瑞虎9', '瑞虎9X', '全新一代瑞虎9'],
    '瑞虎9 C-DM':    ['瑞虎9 C-DM', '瑞虎9高性能版'],
    'QQ冰淇淋':      ['冰淇淋'],
    'QQ3 EV':        ['QQ3'],
    '小蚂蚁':        ['小蚂蚁'],
    '风云A8':        ['风云A8'],
    '风云A8L':       ['风云A8L'],
    '风云A9L':       ['风云A9L'],
    '风云T8':        ['风云T8'],
    '风云T9':        ['风云T9'],
    '风云T9L':       ['风云T9L'],
    '风云T10':       ['2025款 风云T10'],
    '风云T11':       ['风云T11'],
    '风云X3':        ['风云X3'],
    '风云X3 PLUS':   ['风云X3 PLUS'],
    '风云X3L':       ['风云X3L'],
    # 标准库无对应（直接标黄）：艾瑞泽5 GT / 艾瑞泽5 PLUS / 风云T6 / 瑞虎8 PRO新能源
}


def norm(s):
    if s is None:
        return ''
    return re.sub(r'\s+', '', str(s)).replace('（', '(').replace('）', ')').upper()


def num(x):
    if x is None:
        return None
    m = re.search(r'([\d.]+)', str(x))
    return float(m.group(1)) if m else None


def wan_to_yuan(s):
    v = num(s)
    return None if v is None else round(v * 10000)


def cjk_overlap(a, b):
    """两名称的字符集合 Jaccard 相似度，用于同价多车型消歧。"""
    sa, sb = set(a), set(b)
    if not sa or not sb:
        return 0.0
    return len(sa & sb) / len(sa | sb)


# ── 读取标准库 ──────────────────────────────────────────────────────────
def load_standard():
    wb = openpyxl.load_workbook(STD_PATH)
    ws = wb['网销平台报价建议表']
    merged = {}
    for mc in ws.merged_cells.ranges:
        tl = ws.cell(mc.min_row, mc.min_col).value
        for r in range(mc.min_row, mc.max_row + 1):
            for c in range(mc.min_col, mc.max_col + 1):
                merged[(r, c)] = tl

    def gv(r, c):
        return merged.get((r, c), ws.cell(r, c).value)

    by_series = collections.defaultdict(list)
    for r in range(4, 200):
        media = gv(r, 6)
        if not media:
            continue
        mkt = num(gv(r, 5))
        if mkt is None:
            continue
        disc = num(gv(r, 8)) or 0
        by_series[str(gv(r, 2))].append({
            'media': str(media),
            'media_norm': norm(media),
            'mkt': round(mkt),
            'bare': round(mkt - disc),
        })
    return by_series


def match_std_bare(our_series, our_trim, our_msrp, std):
    """返回 (std_bare or None, series_matched: bool)。"""
    cand_series = SERIES_MAP.get(our_series)
    if not cand_series:
        return None, False
    cands = []
    for cs in cand_series:
        cands.extend(std.get(cs, []))
    if not cands:
        return None, False

    # 用指导价(=市场价格)定位
    m = [c for c in cands if our_msrp is not None and c['mkt'] == our_msrp]
    if not m:
        return None, True          # 车系在标准内，但该车型(指导价)对不上
    bares = set(c['bare'] for c in m)
    if len(bares) == 1:
        return m[0]['bare'], True
    # 同价多车型：用名称字符重合度消歧
    tn = norm(our_trim)
    best = max(m, key=lambda c: cjk_overlap(tn, c['media_norm']))
    return best['bare'], True


def main():
    std = load_standard()
    wb = openpyxl.load_workbook(SRC_PATH)
    ws = wb.active

    n_red = n_yellow = n_ok = 0
    cache = {}
    red_examples = []
    red_rows = []          # 全部标红明细
    yellow_combos = collections.Counter()

    for r in range(2, ws.max_row + 1):
        series = ws.cell(r, 5).value
        trim = ws.cell(r, 6).value
        bare_cell = ws.cell(r, 7)
        msrp = wan_to_yuan(ws.cell(r, 8).value)
        if not trim:
            continue
        key = (series, trim, msrp)
        if key not in cache:
            cache[key] = match_std_bare(series, trim, msrp, std)
        std_bare, series_matched = cache[key]

        if std_bare is None:
            ws.cell(r, 5).fill = YELLOW
            n_yellow += 1
            yellow_combos[(series, trim)] += 1
            continue

        our_bare = wan_to_yuan(bare_cell.value)
        if our_bare != std_bare:
            bare_cell.fill = RED
            n_red += 1
            row_info = (ws.cell(r, 1).value, ws.cell(r, 2).value,
                        ws.cell(r, 3).value, ws.cell(r, 4).value, series, trim,
                        bare_cell.value, f'{std_bare/10000:.2f}万')
            red_rows.append(row_info)
            if len(red_examples) < 25:
                red_examples.append((ws.cell(r, 4).value, series, trim,
                                     bare_cell.value, f'{std_bare/10000:.2f}万'))
        else:
            n_ok += 1

    # ── 汇总说明 sheet ─────────────────────────────────────────────────
    from openpyxl.styles import Font, Alignment
    s2 = wb.create_sheet('核对结果汇总')
    bold = Font(bold=True)
    s2.append(['汽车之家 奇瑞经销商裸车价 — 网销平台报价标准(2026.6.2)核对结果'])
    s2['A1'].font = Font(bold=True, size=13)
    s2.append([])
    s2.append(['图例'])
    s2['A3'].font = bold
    c = s2.cell(4, 1, '裸车价与标准不符（0容错）— 标红'); c.fill = RED
    c = s2.cell(5, 1, '该车型未在 2026.6.2 标准内，无法核对 — 车系标黄'); c.fill = YELLOW
    s2.append([])
    s2.append(['统计', '行数']); s2['A7'].font = bold; s2['B7'].font = bold
    s2.append(['车型行总数', n_ok + n_red + n_yellow])
    s2.append(['裸车价相符', n_ok])
    s2.append(['裸车价不符(标红)', n_red])
    s2.append(['无法匹配标准(标黄)', n_yellow])
    s2.append([])
    hdr_row = s2.max_row + 1
    s2.append(['【标红明细】省份', '城市', '经销商ID', '经销商', '车系', '车型',
               '我方裸车价', '标准裸车价'])
    for cidx in range(1, 9):
        s2.cell(hdr_row, cidx).font = bold
    for info in red_rows:
        s2.append(list(info))
    # 列宽
    for col, w in zip('ABCDEFGH', (16, 10, 12, 26, 14, 40, 12, 12)):
        s2.column_dimensions[col].width = w

    wb.save(OUT_PATH)
    total = n_ok + n_red + n_yellow
    print(f'输出: {OUT_PATH}')
    print(f'  总车型行: {total}')
    print(f'  ✅ 裸车价相符        : {n_ok}')
    print(f'  🟥 裸车价不符(标红)  : {n_red}')
    print(f'  🟨 无法匹配标准(标黄): {n_yellow}')

    print('\n标红示例(经销商 | 车系 | 车型 | 我方裸车价 → 标准裸车价):')
    for d, s, t, ob, sb in red_examples:
        print(f'   {d} | {s} | {t} | {ob} → {sb}')

    print('\n标黄(无法匹配标准)车型清单:')
    for (s, t), c in yellow_combos.most_common():
        print(f'   [{c:>3}家] {s} | {t}')


if __name__ == '__main__':
    main()
