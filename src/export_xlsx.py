"""
Step 4 — 输出 .xlsx（含内嵌头图截图）

布局：一「车型」一行；经销商级字段（省/市/ID/经销商名称/最新软文日期/头图截图）
在同一经销商的多行间纵向合并居中，截图只插一次。
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ── 样式 ──────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="003087")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 图片嵌入参数
IMG_W, IMG_H = 320, 142          # 显示像素
IMG_COL_WIDTH = 46               # 字符宽
IMG_ROW_HEIGHT = 26              # 每行磅高（trim 行）

# ── 列定义：(header, width, kind) ─────────────────────────────────────────
# kind: "dealer"=经销商级(纵向合并)  "trim"=车型级  "img"=截图(特殊)
def _build_columns(features: dict) -> list[tuple]:
    cols = [
        ("省份",       8,  "dealer"),
        ("城市",       8,  "dealer"),
        ("经销商ID",   11, "dealer"),
        ("经销商名称", 24, "dealer"),
    ]
    if features.get("pricing"):
        cols += [
            ("车系",   12, "trim"),
            ("车型",   34, "trim"),
            ("裸车价",  9, "trim"),
            ("指导价",  9, "trim"),
        ]
    if features.get("article_date"):
        cols.append(("最新软文日期", 14, "dealer"))
    if features.get("screenshot"):
        cols.append(("头图截图", IMG_COL_WIDTH, "img"))
    return cols


def _dealer_value(header: str, rec: dict) -> str:
    return {
        "省份": rec.get("province", ""),
        "城市": rec.get("city", ""),
        "经销商ID": rec.get("dealer_id", ""),
        "经销商名称": rec.get("name", ""),
        "最新软文日期": rec.get("article_date", ""),
    }.get(header, "")


def export_to_xlsx(records: list[dict], output_path: str, features: dict):
    """
    records: 每项为经销商详情 {province,city,dealer_id,name,pricing[],article_date,screenshot_path}
    pricing 每项: {series,trim,price_bare,price_msrp}
    """
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "奇瑞经销商报价"
    ws.freeze_panes = "A2"

    cols = _build_columns(features)
    has_pricing = features.get("pricing")
    has_img = features.get("screenshot")
    img_col_idx = next((i for i, c in enumerate(cols, 1) if c[2] == "img"), None)

    # 表头
    for ci, (header, width, _) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=header)
        cell.font, cell.fill, cell.alignment, cell.border = HEADER_FONT, HEADER_FILL, CENTER, BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 26

    # 数据
    row = 2
    for rec in records:
        # 该经销商的车型行（pricing 关闭或为空时占 1 行）
        trims = rec.get("pricing") or [] if has_pricing else []
        if has_pricing and not trims:
            trims = [{}]            # 无报价也保留 1 行
        n = len(trims) if has_pricing else 1
        start_row = row

        for k in range(n):
            trim = trims[k] if has_pricing else {}
            for ci, (header, _, kind) in enumerate(cols, 1):
                cell = ws.cell(row=row, column=ci)
                cell.border = BORDER
                if kind == "dealer":
                    cell.alignment = LEFT if header in ("经销商名称",) else CENTER
                    cell.value = _dealer_value(header, rec) if k == 0 else None
                elif kind == "img":
                    cell.alignment = CENTER
                    cell.value = None
                elif kind == "trim":
                    cell.alignment = LEFT if header == "车型" else CENTER
                    cell.value = {
                        "车系": trim.get("series", ""),
                        "车型": trim.get("trim", ""),
                        "裸车价": trim.get("price_bare", ""),
                        "指导价": trim.get("price_msrp", ""),
                    }.get(header, "")
            ws.row_dimensions[row].height = max(IMG_ROW_HEIGHT, 18)
            row += 1

        end_row = row - 1

        # 纵向合并经销商级列 + 截图列
        if end_row > start_row:
            for ci, (header, _, kind) in enumerate(cols, 1):
                if kind in ("dealer", "img"):
                    ws.merge_cells(
                        start_row=start_row, start_column=ci,
                        end_row=end_row, end_column=ci,
                    )
                    mc = ws.cell(row=start_row, column=ci)
                    mc.alignment = LEFT if header == "经销商名称" else CENTER

        # 插入截图（合并区左上角）
        if has_img and img_col_idx:
            p = rec.get("screenshot_path", "")
            if p and Path(p).exists():
                try:
                    im = XLImage(p)
                    im.width, im.height = IMG_W, IMG_H
                    ws.add_image(im, f"{get_column_letter(img_col_idx)}{start_row}")
                    # 保证行高容纳图片
                    span = end_row - start_row + 1
                    need = IMG_H * 0.75 / span + 2   # 磅
                    for r in range(start_row, end_row + 1):
                        if (ws.row_dimensions[r].height or 0) < need:
                            ws.row_dimensions[r].height = need
                except Exception as e:
                    ws.cell(row=start_row, column=img_col_idx).value = f"[图片错误:{e}]"

    wb.save(output_path)
    n_rows = row - 2
    print(f"[XLSX] 已保存: {output_path}  ({len(records)} 家经销商 / {n_rows} 行车型)")
