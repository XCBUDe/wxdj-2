#!/usr/bin/env python3
"""
汽车之家奇瑞经销商点检 — 一键启动器

用法：
  1. 修改下方 TASK_CONFIG 中的文件路径（dealer_list_xlsx / standard_xlsx）
  2. python task.py

完整流程：
  名单 Excel → 批量 Playwright 采集 → result.xlsx → 与报价标准对比 → result_checked.xlsx + 截图.zip

检查点续跑：中途中断后直接重跑，已完成的经销商自动跳过。
"""
# =============================================================================
# ★ 定制区：只需改这里 ★
# =============================================================================
TASK_CONFIG = {
    # ── 输入 ─────────────────────────────────────────────────────────────
    # 留空即可！程序会自动识别本文件夹里的 xlsx：
    #   文件名含「检核 / 名单 / uid」→ 当作经销商名单
    #   文件名含「报价 / 标准」      → 当作报价标准
    # 如需手动指定，把文件名填进下面引号里（不要删掉键名）。
    "dealer_list_xlsx": "",   # 经销商名单（留空=自动识别）
    "standard_xlsx":    "",   # 报价标准  （留空=自动识别 / 不核价就留空）

    # ── 输出 ─────────────────────────────────────────────────────────────
    "output_dir":       "output",
    "result_xlsx":      "output/result.xlsx",
    "checked_xlsx":     "output/result_checked.xlsx",   # 报价对比着色结果
    "screenshots_zip":  "output/头图截图.zip",

    # ── 断点续跑 ─────────────────────────────────────────────────────────
    "checkpoint":       "output/checkpoint.jsonl",

    # ── 采集开关 ─────────────────────────────────────────────────────────
    "concurrency": 2,   # 并发浏览器数（建议 1-3，超过 4 易触发反爬）
    "features": {
        "pricing":      True,   # 采集车型报价
        "screenshot":   True,   # 采集头图截图
        "article_date": True,   # 采集最新软文日期
    },

    # ── 后处理 ───────────────────────────────────────────────────────────
    "compare_prices":   True,   # 执行报价对比着色（需 standard_xlsx）
    "zip_screenshots":  True,   # 打包截图 ZIP（按经销商简称命名）
}
# =============================================================================

import re
import sys
import time
import json
import zipfile
import collections
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl
from openpyxl.styles import PatternFill


# ── 省/城市推断（从简称关键字推断） ─────────────────────────────────────
_PROV_CITY = [
    ("福建", ["福清"],         "福州"),
    ("福建", ["福建", "厦门"], "厦门"),
    ("福建", ["龙岩"],         "龙岩"),
    ("福建", ["泉州", "晋江"], "泉州"),
    ("福建", ["三明"],         "三明"),
    ("福建", ["漳州"],         "漳州"),
    ("广东", ["东莞"],         "东莞"),
    ("广东", ["佛山"],         "佛山"),
    ("广东", ["广州", "广东"], "广州"),
    ("广东", ["惠州"],         "惠州"),
    ("广东", ["江门"],         "江门"),
    ("广东", ["揭阳"],         "揭阳"),
    ("广东", ["茂名"],         "茂名"),
    ("广东", ["梅州"],         "梅州"),
    ("广东", ["清远"],         "清远"),
    ("广东", ["汕头"],         "汕头"),
    ("广东", ["韶关"],         "韶关"),
    ("广东", ["深圳"],         "深圳"),
    ("广东", ["湛江"],         "湛江"),
    ("广东", ["肇庆"],         "肇庆"),
    ("广东", ["中山"],         "中山"),
    ("广东", ["珠海"],         "珠海"),
    ("海南", ["海南"],         "海口"),
]


def _infer_prov_city(short_name: str):
    for prov, kws, city in _PROV_CITY:
        if any(k in short_name for k in kws):
            return prov, city
    return "未知", "未知"


# ── Step 1: 从 Excel 名单加载经销商 ─────────────────────────────────────
def load_dealers(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Sheet1"]
    dealers = []
    for r in range(4, ws.max_row + 1):
        uid   = ws.cell(r, 3).value
        short = ws.cell(r, 2).value or ""
        if uid is None or str(uid) == "#N/A" or not str(uid).isdigit():
            continue
        prov, city = _infer_prov_city(short)
        dealers.append({
            "province":  prov,
            "city":      city,
            "dealer_id": str(uid),
            "name":      short,
        })
    return dealers


# ── Step 2: 采集（含断点续跑） ───────────────────────────────────────────
def _load_checkpoint(path: str) -> set[str]:
    done = set()
    if Path(path).exists():
        with open(path, encoding="utf-8") as f:
            for line in f:
                try:
                    done.add(json.loads(line.strip())["dealer_id"])
                except Exception:
                    pass
    return done


def _append_checkpoint(path: str, record: dict):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with open(path, "a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False) + "\n")


def _load_all_checkpoint(path: str) -> list[dict]:
    records = []
    if Path(path).exists():
        with open(path, encoding="utf-8") as f:
            for line in f:
                try:
                    records.append(json.loads(line.strip()))
                except Exception:
                    pass
    return records


def run_scrape(dealers: list[dict], cfg: dict) -> list[dict]:
    from src.detail import fetch_dealer_detail
    from src.export_xlsx import export_to_xlsx

    features    = cfg["features"]
    checkpoint  = cfg["checkpoint"]
    concurrency = cfg["concurrency"]
    result_path = cfg["result_xlsx"]

    Path(result_path).parent.mkdir(parents=True, exist_ok=True)
    done_ids = _load_checkpoint(checkpoint)
    todo     = [d for d in dealers if d["dealer_id"] not in done_ids]
    already  = _load_all_checkpoint(checkpoint)

    print(f"[采集] 名单共 {len(dealers)} 家  待采集 {len(todo)} 家  已完成 {len(already)} 家")

    def _work(dealer: dict) -> dict:
        result = fetch_dealer_detail(dealer, features, browser=None)
        _append_checkpoint(checkpoint, result)
        time.sleep(1.5)
        return result

    new_records = []
    with ThreadPoolExecutor(max_workers=concurrency) as pool:
        futures = {pool.submit(_work, d): d for d in todo}
        for i, fut in enumerate(as_completed(futures), 1):
            d = futures[fut]
            try:
                rec = fut.result()
                new_records.append(rec)
                n = len(rec.get("pricing") or [])
                st = f"OK {n}款" if not rec.get("error") else f"ERR {rec['error'][:40]}"
                print(f"  [{i:03d}/{len(todo)}] {d['dealer_id']} {rec.get('name') or d['name']} → {st}")
            except Exception as e:
                print(f"  [{i:03d}/{len(todo)}] {d['dealer_id']} FATAL: {e}")

    all_records = already + new_records
    export_to_xlsx(all_records, result_path, features)
    print(f"[采集] 完成！共 {len(all_records)} 家 → {result_path}")
    return all_records


# ── Step 3: 报价对比着色 ─────────────────────────────────────────────────

# 我方车系 → 标准车系候选映射
_SERIES_MAP = {
    '艾瑞泽5':       ['艾瑞泽5 卓越版'],
    '艾瑞泽8':       ['2025款 艾瑞泽8', '2025款艾瑞泽8卓越版'],
    '艾瑞泽8 PRO':   ['艾瑞泽8 PRO', '艾瑞泽8 PRO 400T'],
    '瑞虎3x':        ['瑞虎3x 卓越版'],
    '瑞虎5':         ['瑞虎5'],
    '瑞虎5x':        ['瑞虎5x高能版', '瑞虎5x 卓越版'],
    '瑞虎7':         ['瑞虎7卓越版\n（全新一代瑞虎7）', '瑞虎7 高能版', '全新瑞虎7', '全新瑞虎7 C-DM'],
    '瑞虎7 PLUS':    ['瑞虎7 PLUS'],
    '瑞虎7L':        ['瑞虎7L'],
    '瑞虎8':         ['第五代瑞虎8', '瑞虎8 卓越版'],
    '瑞虎8 PLUS':    ['全新瑞虎8 PLUS'],
    '瑞虎8 PLUS C-DM': ['全新瑞虎8 PLUS C-DM'],
    '瑞虎8 PRO':     ['瑞虎8 PRO'],
    '瑞虎8L':        ['瑞虎8 L'],
    '瑞虎9':         ['瑞虎9', '瑞虎9X', '全新一代瑞虎9'],
    '瑞虎9 C-DM':    ['瑞虎9 C-DM', '瑞虎9高性能版'],
    'QQ冰淇淋':      ['冰淇淋'],
    'QQ3 EV':        ['QQ3'],
    '小蚂蚁':        ['小蚂蚁'],
    '风云A8':        ['风云A8'], '风云A8L': ['风云A8L'], '风云A9L': ['风云A9L'],
    '风云T8':        ['风云T8'], '风云T9':  ['风云T9'],  '风云T9L': ['风云T9L'],
    '风云T10':       ['2025款 风云T10'], '风云T11': ['风云T11'],
    '风云X3':        ['风云X3'], '风云X3 PLUS': ['风云X3 PLUS'], '风云X3L': ['风云X3L'],
}
RED    = PatternFill('solid', fgColor='FFC7CE')
YELLOW = PatternFill('solid', fgColor='FFEB9C')


def _num(x):
    if x is None:
        return None
    m = re.search(r'([\d.]+)', str(x))
    return float(m.group(1)) if m else None


def _wan2yuan(s):
    v = _num(s)
    return None if v is None else round(v * 10000)


def _norm(s):
    return re.sub(r'\s+', '', str(s or '')).replace('（', '(').replace('）', ')').upper()


def _cjk_overlap(a, b):
    sa, sb = set(a), set(b)
    return len(sa & sb) / len(sa | sb) if sa and sb else 0.0


def _load_standard(xlsx_path: str) -> dict:
    """返回 {std_series: [{media, media_norm, mkt, bare}]}"""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb['网销平台报价建议表']
    merged = {}
    for mc in ws.merged_cells.ranges:
        tl = ws.cell(mc.min_row, mc.min_col).value
        for r in range(mc.min_row, mc.max_row + 1):
            for c in range(mc.min_col, mc.max_col + 1):
                merged[(r, c)] = tl

    def gv(r, c):
        return merged.get((r, c), ws.cell(r, c).value)

    by_series = collections.defaultdict(list)
    for r in range(4, 220):
        media = gv(r, 6)
        if not media:
            continue
        mkt = _num(gv(r, 5))
        if mkt is None:
            continue
        disc = _num(gv(r, 8)) or 0
        by_series[str(gv(r, 2))].append({
            'media': str(media),
            'media_norm': _norm(media),
            'mkt': round(mkt),
            'bare': round(mkt - disc),
        })
    return by_series


def _match_bare(our_series, our_trim, our_msrp, std):
    cs_names = _SERIES_MAP.get(our_series)
    if not cs_names:
        return None, False
    cands = []
    for cs in cs_names:
        cands.extend(std.get(cs, []))
    if not cands:
        return None, False
    m = [c for c in cands if our_msrp is not None and c['mkt'] == our_msrp]
    if not m:
        return None, True
    bares = set(c['bare'] for c in m)
    if len(bares) == 1:
        return m[0]['bare'], True
    tn = _norm(our_trim)
    best = max(m, key=lambda c: _cjk_overlap(tn, c['media_norm']))
    return best['bare'], True


def run_compare(result_xlsx: str, standard_xlsx: str, checked_xlsx: str):
    """对比报价标准，输出着色 xlsx。"""
    from openpyxl.styles import Font

    std = _load_standard(standard_xlsx)
    wb  = openpyxl.load_workbook(result_xlsx)
    ws  = wb.active

    n_ok = n_red = n_yellow = 0
    cache = {}
    red_rows = []
    yellow_combos = collections.Counter()

    for r in range(2, ws.max_row + 1):
        series    = ws.cell(r, 5).value
        trim      = ws.cell(r, 6).value
        bare_cell = ws.cell(r, 7)
        msrp      = _wan2yuan(ws.cell(r, 8).value)
        if not trim:
            continue
        key = (series, trim, msrp)
        if key not in cache:
            cache[key] = _match_bare(series, trim, msrp, std)
        std_bare, series_matched = cache[key]

        if std_bare is None:
            ws.cell(r, 5).fill = YELLOW
            n_yellow += 1
            yellow_combos[(series, trim)] += 1
        elif _wan2yuan(bare_cell.value) != std_bare:
            bare_cell.fill = RED
            n_red += 1
            red_rows.append((
                ws.cell(r, 1).value, ws.cell(r, 2).value,
                ws.cell(r, 3).value, ws.cell(r, 4).value,
                series, trim, bare_cell.value, f'{std_bare / 10000:.2f}万',
            ))
        else:
            n_ok += 1

    # 汇总 sheet
    s2 = wb.create_sheet('核对结果汇总')
    bold = Font(bold=True)
    s2.append(['报价标准核对结果'])
    s2['A1'].font = Font(bold=True, size=13)
    s2.append([])
    s2.append(['图例'])
    s2['A3'].font = bold
    s2.cell(4, 1, '裸车价与标准不符（0容错）— 标红').fill = RED
    s2.cell(5, 1, '该车型不在 2026.6.2 标准内，无法核对 — 车系标黄').fill = YELLOW
    s2.append([])
    for label, val in [('车型行总数', n_ok + n_red + n_yellow),
                       ('裸车价相符', n_ok),
                       ('裸车价不符(标红)', n_red),
                       ('无法匹配标准(标黄)', n_yellow)]:
        r_obj = s2.append([label, val])
    s2.append([])
    hdr = s2.max_row + 1
    s2.append(['省份', '城市', '经销商ID', '经销商名称', '车系', '车型', '采集裸车价', '标准裸车价'])
    for ci in range(1, 9):
        s2.cell(hdr, ci).font = bold
    for row in red_rows:
        s2.append(list(row))
    for col, w in zip('ABCDEFGH', (12, 10, 12, 28, 16, 44, 12, 12)):
        s2.column_dimensions[col].width = w

    wb.save(checked_xlsx)
    total = n_ok + n_red + n_yellow
    print(f"[核价] 完成！共 {total} 行  相符 {n_ok}  标红 {n_red}  标黄 {n_yellow} → {checked_xlsx}")
    return n_red, n_yellow


# ── Step 4: 打包截图 ZIP ────────────────────────────────────────────────
def run_zip_screenshots(dealers: list[dict], cfg: dict):
    src = Path(cfg["output_dir"]) / "screenshots"
    out = cfg["screenshots_zip"]
    uid_to_short = {d["dealer_id"]: d["name"] for d in dealers}
    missing = []

    with zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED) as zf:
        for png in sorted(src.glob("*.png")):
            if "test" in png.name:
                continue
            did   = png.stem
            short = uid_to_short.get(did, did)
            zf.write(png, arcname=f"{short}.png")
        # 找无截图的经销商
        shot_ids = {p.stem for p in src.glob("*.png")}
        for d in dealers:
            if d["dealer_id"] not in shot_ids:
                missing.append(f"{d['name']}({d['dealer_id']})")
        if missing:
            note = "以下经销商无截图（采集时被反爬拦截或无法访问）：\n" + "\n".join(missing) + "\n"
            zf.writestr("【缺失说明】.txt", note)

    print(f"[截图] ZIP 已保存: {out}")
    if missing:
        print(f"  无截图经销商({len(missing)}家): {', '.join(missing)}")


def _autodetect_inputs(cfg):
    """名单/标准路径为空或不存在时，自动扫描脚本所在文件夹里的 xlsx。"""
    here = Path(__file__).resolve().parent
    candidates = [
        p for p in here.glob("*.xlsx")
        if not p.name.startswith("~$")        # 排除 Excel 临时锁文件
        and "result" not in p.name.lower()    # 排除我们自己的输出
    ]

    def _resolve(key, keywords):
        val = (cfg.get(key) or "").strip()
        if val and Path(val).exists():
            return                            # 用户已填且文件存在
        for p in candidates:
            if any(kw in p.name for kw in keywords):
                cfg[key] = str(p)
                print(f"[自动识别] {key} → {p.name}")
                return

    _resolve("dealer_list_xlsx", ["检核", "名单", "uid", "经销商"])
    _resolve("standard_xlsx",    ["报价", "标准", "建议"])


# ── 主流程 ──────────────────────────────────────────────────────────────
def main():
    cfg = TASK_CONFIG
    _autodetect_inputs(cfg)
    if not (cfg.get("dealer_list_xlsx") and Path(cfg["dealer_list_xlsx"]).exists()):
        print("\n[错误] 没找到经销商名单 xlsx。")
        print("       请把名单文件（文件名建议含『检核』或『名单』）放到本文件夹后重跑。")
        return
    Path(cfg["output_dir"]).mkdir(parents=True, exist_ok=True)

    print("=" * 60)
    print("  汽车之家经销商点检任务")
    print("=" * 60)
    print(f"  名单:    {cfg['dealer_list_xlsx']}")
    print(f"  标准:    {cfg['standard_xlsx'] or '（未配置，跳过报价对比）'}")
    print(f"  并发数:  {cfg['concurrency']}")
    print(f"  功能:    { {k for k,v in cfg['features'].items() if v} }")
    print("=" * 60)

    # 1. 加载名单
    dealers = load_dealers(cfg["dealer_list_xlsx"])
    print(f"[名单] 共 {len(dealers)} 家经销商")

    # 2. 采集
    run_scrape(dealers, cfg)

    # 3. 报价对比
    if cfg.get("compare_prices") and cfg.get("standard_xlsx"):
        run_compare(cfg["result_xlsx"], cfg["standard_xlsx"], cfg["checked_xlsx"])
    elif cfg.get("compare_prices"):
        print("[核价] standard_xlsx 未配置，跳过报价对比")

    # 4. 打包截图
    if cfg.get("zip_screenshots") and cfg["features"].get("screenshot"):
        run_zip_screenshots(dealers, cfg)

    print("\n✅ 全部完成")
    out_xlsx = cfg["checked_xlsx"] if cfg.get("compare_prices") and cfg.get("standard_xlsx") else cfg["result_xlsx"]
    print(f"   主结果: {out_xlsx}")
    if cfg.get("zip_screenshots"):
        print(f"   截图包: {cfg['screenshots_zip']}")


if __name__ == "__main__":
    main()
