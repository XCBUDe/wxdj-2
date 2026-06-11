#!/usr/bin/env python3
"""
易车奇瑞经销商点检 — 一键启动器

用法：
  1. 把经销商名单 xlsx 放到本文件夹（或在 TASK_CONFIG 填路径）
  2. python yiche_task.py

输入名单格式（表头自动识别，插行插列不影响）：
  含「易车UID」「易车uid」「易车id」列  → 经销商UID
  含「简称」列                          → 经销商简称
  含「省」列                            → 省份
  含「城市」「市」列                     → 城市

输出：
  output_yiche/result.xlsx         ← 采集结果
  output_yiche/result_checked.xlsx ← 报价对比着色（红=不符，黄=无法匹配）
  output_yiche/头图截图.zip        ← 头图 PNG 按简称命名
"""
# =============================================================================
# ★ 定制区：只需改这里（全部留空 = 自动识别）★
# =============================================================================
TASK_CONFIG = {
    # 输入（留空=自动识别本文件夹里的 xlsx）
    "dealer_list_xlsx": "",   # 经销商名单
    "standard_xlsx":    "",   # 报价标准（留空=自动识别；不核价就填 None）

    # 输出目录
    "output_dir":       "output_yiche",
    "result_xlsx":      "output_yiche/result.xlsx",
    "checked_xlsx":     "output_yiche/result_checked.xlsx",
    "screenshots_zip":  "output_yiche/头图截图.zip",
    "checkpoint":       "output_yiche/checkpoint.jsonl",

    # 采集开关
    "concurrency": 3,          # 并发请求数（易车无需浏览器，3-5 问题不大）
    "features": {
        "pricing":      True,
        "article_date": True,
        "screenshot":   True,
    },

    # 后处理
    "compare_prices":  True,   # 执行报价对比着色
    "zip_screenshots": True,   # 打包截图 ZIP
}
# =============================================================================

import re, sys, time, json, zipfile, collections
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl
from openpyxl.styles import PatternFill

RED    = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")


# ─────────────────────────────────────────────────────────────────────────────
# 工具函数
# ─────────────────────────────────────────────────────────────────────────────

def _num(x):
    if x is None:
        return None
    m = re.search(r"([\d.]+)", str(x))
    return float(m.group(1)) if m else None


def _wan2yuan(s):
    v = _num(s)
    return None if v is None else round(v * 10000)


def _norm(s):
    return re.sub(r"\s+", "", str(s or "")).replace("（", "(").replace("）", ")").upper()


def _cjk_overlap(a, b):
    sa, sb = set(a), set(b)
    return len(sa & sb) / len(sa | sb) if sa and sb else 0.0


# ─────────────────────────────────────────────────────────────────────────────
# 自动检测输入文件（按表头关键词定位 UID 列等）
# ─────────────────────────────────────────────────────────────────────────────

def _autodetect_inputs(cfg):
    """名单/标准路径为空时，自动扫描脚本所在文件夹里的 xlsx。"""
    here = Path(__file__).resolve().parent
    candidates = [
        p for p in here.glob("*.xlsx")
        if not p.name.startswith("~$")
        and "result" not in p.name.lower()
        and "output" not in str(p.parent).lower()
    ]

    def _resolve(key, keywords):
        val = (cfg.get(key) or "").strip()
        if val and Path(val).exists():
            return
        for p in candidates:
            if any(kw in p.name for kw in keywords):
                cfg[key] = str(p)
                print(f"[自动识别] {key} → {p.name}")
                return

    _resolve("dealer_list_xlsx", ["检核", "名单", "uid", "UID", "经销商"])
    _resolve("standard_xlsx",    ["报价", "标准", "建议"])


def _find_column(ws, header_row: int, *keywords) -> int | None:
    """在指定行扫描表头，返回第一个含任意关键词的列号（1-based），找不到返回 None。"""
    for col in range(1, ws.max_column + 1):
        v = str(ws.cell(header_row, col).value or "")
        if any(kw in v for kw in keywords):
            return col
    return None


def _find_header_row(ws, uid_keywords=("易车UID", "易车uid", "易车ID", "易车id")) -> int:
    """扫描前10行找含易车UID关键词的表头行，返回行号。"""
    for r in range(1, 11):
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(r, c).value or "")
            if any(kw in v for kw in uid_keywords):
                return r
    return 1  # 默认第1行


# ─────────────────────────────────────────────────────────────────────────────
# 加载经销商名单（表头自动定位，防插行插列）
# ─────────────────────────────────────────────────────────────────────────────

def load_dealers_yiche(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # 找表头行
    header_row = _find_header_row(ws)

    # 找各列
    uid_col  = _find_column(ws, header_row, "易车UID", "易车uid", "易车ID", "易车id")
    name_col = _find_column(ws, header_row, "简称")
    prov_col = _find_column(ws, header_row, "省份", "省")
    city_col = _find_column(ws, header_row, "城市", "市")

    if uid_col is None:
        raise ValueError(
            f"在 {xlsx_path} 的前10行找不到「易车UID」列。"
            "请确认表头含『易车UID』字样。"
        )

    dealers = []
    for r in range(header_row + 1, ws.max_row + 1):
        uid = ws.cell(r, uid_col).value
        if uid is None or str(uid).strip() in ("", "无", "#N/A"):
            continue
        uid = str(uid).strip()
        # 易车 UID 格式：9位数字，100XXXXXX
        if not re.match(r"^1\d{8}$", uid):
            continue
        name = str(ws.cell(r, name_col).value or "") if name_col else ""
        prov = str(ws.cell(r, prov_col).value or "") if prov_col else ""
        city = str(ws.cell(r, city_col).value or "") if city_col else ""
        dealers.append({
            "province":  prov or "未知",
            "city":      city or "未知",
            "dealer_id": uid,
            "name":      name,
        })
    return dealers


# ─────────────────────────────────────────────────────────────────────────────
# 断点续跑
# ─────────────────────────────────────────────────────────────────────────────

def _load_checkpoint(path) -> set:
    done = set()
    if Path(path).exists():
        with open(path, encoding="utf-8") as f:
            for line in f:
                try:
                    done.add(json.loads(line)["dealer_id"])
                except Exception:
                    pass
    return done


def _append_checkpoint(path, rec):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with open(path, "a", encoding="utf-8") as f:
        f.write(json.dumps(rec, ensure_ascii=False) + "\n")


def _load_all_checkpoint(path) -> list:
    recs = []
    if Path(path).exists():
        with open(path, encoding="utf-8") as f:
            for line in f:
                try:
                    recs.append(json.loads(line))
                except Exception:
                    pass
    return recs


# ─────────────────────────────────────────────────────────────────────────────
# 批量采集
# ─────────────────────────────────────────────────────────────────────────────

def run_scrape(dealers, cfg) -> list:
    from src.yiche_detail import fetch_dealer_yiche
    from src.export_xlsx import export_to_xlsx

    features    = cfg["features"]
    checkpoint  = cfg["checkpoint"]
    concurrency = cfg["concurrency"]
    result_path = cfg["result_xlsx"]
    scr_dir     = str(Path(cfg["output_dir"]) / "screenshots_yiche")

    Path(result_path).parent.mkdir(parents=True, exist_ok=True)
    done_ids = _load_checkpoint(checkpoint)
    todo     = [d for d in dealers if d["dealer_id"] not in done_ids]
    already  = _load_all_checkpoint(checkpoint)

    print(f"[采集] 名单共 {len(dealers)} 家  待采集 {len(todo)} 家  已完成 {len(already)} 家")

    def _work(dealer):
        rec = fetch_dealer_yiche(dealer, features, screenshot_dir=scr_dir)
        _append_checkpoint(checkpoint, rec)
        return rec

    new_recs = []
    with ThreadPoolExecutor(max_workers=concurrency) as pool:
        futures = {pool.submit(_work, d): d for d in todo}
        for i, fut in enumerate(as_completed(futures), 1):
            d = futures[fut]
            try:
                rec = fut.result()
                new_recs.append(rec)
                n  = len(rec.get("pricing") or [])
                st = f"OK {n}款" if not rec.get("error") else f"ERR {rec['error'][:50]}"
                print(f"  [{i:03d}/{len(todo)}] {d['dealer_id']} {rec.get('name') or d['name']} → {st}")
            except Exception as e:
                print(f"  [{i:03d}/{len(todo)}] {d['dealer_id']} FATAL: {e}")

    all_recs = already + new_recs
    export_to_xlsx(all_recs, result_path, features)
    print(f"[采集] 完成！共 {len(all_recs)} 家 → {result_path}")
    return all_recs


# ─────────────────────────────────────────────────────────────────────────────
# 报价对比着色（复用 task.py 里的 SERIES_MAP + 算法）
# ─────────────────────────────────────────────────────────────────────────────

_SERIES_MAP = {
    "艾瑞泽5":       ["艾瑞泽5 卓越版"],
    "艾瑞泽8":       ["2025款 艾瑞泽8", "2025款艾瑞泽8卓越版"],
    "艾瑞泽8 PRO":   ["艾瑞泽8 PRO", "艾瑞泽8 PRO 400T"],
    "瑞虎3x":        ["瑞虎3x 卓越版"],
    "瑞虎5":         ["瑞虎5"],
    "瑞虎5x":        ["瑞虎5x高能版", "瑞虎5x 卓越版"],
    "瑞虎7":         ["瑞虎7卓越版\n（全新一代瑞虎7）", "瑞虎7 高能版", "全新瑞虎7", "全新瑞虎7 C-DM"],
    "瑞虎7 PLUS":    ["瑞虎7 PLUS"],
    "瑞虎7L":        ["瑞虎7L"],
    "瑞虎8":         ["第五代瑞虎8", "瑞虎8 卓越版"],
    "瑞虎8 PLUS":    ["全新瑞虎8 PLUS"],
    "瑞虎8 PLUS C-DM": ["全新瑞虎8 PLUS C-DM"],
    "瑞虎8 PRO":     ["瑞虎8 PRO"],
    "瑞虎8L":        ["瑞虎8 L"],
    "瑞虎9":         ["瑞虎9", "瑞虎9X", "全新一代瑞虎9"],
    "瑞虎9 C-DM":    ["瑞虎9 C-DM", "瑞虎9高性能版"],
    "QQ冰淇淋":      ["冰淇淋"],
    "QQ3 EV":        ["QQ3"],
    "奇瑞QQ3EV":     ["QQ3"],
    "小蚂蚁":        ["小蚂蚁"],
    "风云A8":  ["风云A8"],   "风云A8L": ["风云A8L"], "风云A9L": ["风云A9L"],
    "风云T8":  ["风云T8"],   "风云T9":  ["风云T9"],  "风云T9L": ["风云T9L"],
    "风云T10": ["2025款 风云T10"], "风云T11": ["风云T11"],
    "风云X3":  ["风云X3"],   "风云X3 PLUS": ["风云X3 PLUS"], "风云X3L": ["风云X3L"],
    "奇瑞风云": [],  # 占位，子品牌车系会匹配到具体名字
}


def _load_standard(xlsx_path: str) -> dict:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["网销平台报价建议表"]
    merged = {}
    for mc in ws.merged_cells.ranges:
        tl = ws.cell(mc.min_row, mc.min_col).value
        for r in range(mc.min_row, mc.max_row + 1):
            for c in range(mc.min_col, mc.max_col + 1):
                merged[(r, c)] = tl

    def gv(r, c):
        return merged.get((r, c), ws.cell(r, c).value)

    by_series = collections.defaultdict(list)
    for r in range(4, 220):
        media = gv(r, 6)
        if not media:
            continue
        mkt = _num(gv(r, 5))
        if mkt is None:
            continue
        disc  = _num(gv(r, 8)) or 0
        sname = str(gv(r, 2) or "")
        by_series[sname].append({
            "media":      str(media),
            "media_norm": _norm(media),
            "mkt":        round(mkt),
            "bare":       round(mkt - disc),
        })
    return by_series


def _match_bare(our_series, our_trim, our_msrp, std):
    cs_names = _SERIES_MAP.get(our_series)
    if cs_names is None:
        # 用车系名直接查（易车车系名可能和标准表名一致）
        cs_names = [our_series]
    cands = []
    for cs in cs_names:
        cands.extend(std.get(cs, []))
    if not cands:
        return None, False
    m = [c for c in cands if our_msrp is not None and c["mkt"] == our_msrp]
    if not m:
        return None, True
    bares = set(c["bare"] for c in m)
    if len(bares) == 1:
        return m[0]["bare"], True
    tn = _norm(our_trim)
    best = max(m, key=lambda c: _cjk_overlap(tn, c["media_norm"]))
    return best["bare"], True


def run_compare(result_xlsx, standard_xlsx, checked_xlsx):
    from openpyxl.styles import Font
    std = _load_standard(standard_xlsx)
    wb  = openpyxl.load_workbook(result_xlsx)
    ws  = wb.active

    n_ok = n_red = n_yellow = 0
    cache, red_rows = {}, []
    yellow_combos = collections.Counter()

    for r in range(2, ws.max_row + 1):
        series    = ws.cell(r, 5).value
        trim      = ws.cell(r, 6).value
        bare_cell = ws.cell(r, 7)
        msrp      = _wan2yuan(ws.cell(r, 8).value)
        if not trim:
            continue
        key = (series, trim, msrp)
        if key not in cache:
            cache[key] = _match_bare(series, trim, msrp, std)
        std_bare, series_matched = cache[key]

        if std_bare is None:
            ws.cell(r, 5).fill = YELLOW
            n_yellow += 1
            yellow_combos[(series, trim)] += 1
        elif _wan2yuan(bare_cell.value) != std_bare:
            bare_cell.fill = RED
            n_red += 1
            red_rows.append((
                ws.cell(r, 1).value, ws.cell(r, 2).value,
                ws.cell(r, 3).value, ws.cell(r, 4).value,
                series, trim, bare_cell.value, f"{std_bare/10000:.2f}万",
            ))
        else:
            n_ok += 1

    s2 = wb.create_sheet("核对结果汇总")
    bold = Font(bold=True)
    s2.append(["易车 奇瑞经销商裸车价 — 网销平台报价标准(2026.6.2)核对结果"])
    s2["A1"].font = Font(bold=True, size=13)
    s2.append([])
    s2.cell(4, 1, "裸车价与标准不符（0容错）— 标红").fill = RED
    s2.cell(5, 1, "该车型不在 2026.6.2 标准内，无法核对 — 车系标黄").fill = YELLOW
    s2.append([])
    for label, val in [("车型行总数", n_ok + n_red + n_yellow),
                       ("裸车价相符", n_ok),
                       ("裸车价不符(标红)", n_red),
                       ("无法匹配标准(标黄)", n_yellow)]:
        s2.append([label, val])
    s2.append([])
    hdr = s2.max_row + 1
    s2.append(["省份","城市","经销商ID","经销商名称","车系","车型","采集裸车价","标准裸车价"])
    for ci in range(1, 9):
        s2.cell(hdr, ci).font = bold
    for row in red_rows:
        s2.append(list(row))
    for col, w in zip("ABCDEFGH", (12, 10, 12, 28, 16, 44, 12, 12)):
        s2.column_dimensions[col].width = w

    wb.save(checked_xlsx)
    total = n_ok + n_red + n_yellow
    print(f"[核价] 完成！共 {total} 行  相符 {n_ok}  标红 {n_red}  标黄 {n_yellow} → {checked_xlsx}")
    if red_rows:
        print("  标红示例（经销商 | 车系 | 车型 | 采集裸车价 → 标准裸车价）:")
        for row in red_rows[:10]:
            print(f"    {row[3]} | {row[4]} | {row[5]} | {row[6]} → {row[7]}")
    if yellow_combos:
        print("  标黄车型（无法匹配标准）:")
        for (s, t), cnt in yellow_combos.most_common(10):
            print(f"    [{cnt:>3}家] {s} | {t}")


# ─────────────────────────────────────────────────────────────────────────────
# 截图打包 ZIP
# ─────────────────────────────────────────────────────────────────────────────

def run_zip_screenshots(dealers, cfg):
    scr_dir  = str(Path(cfg["output_dir"]) / "screenshots_yiche")
    zip_path = cfg["screenshots_zip"]
    name_map = {d["dealer_id"]: d["name"] for d in dealers}

    Path(zip_path).parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        missing = []
        for did, sname in name_map.items():
            src = Path(scr_dir) / f"{did}.png"
            if src.exists():
                zf.write(src, f"{sname}.png")
            else:
                missing.append(f"{sname}({did})")
        if missing:
            zf.writestr("【缺失说明】.txt",
                        "以下经销商头图未下载成功：\n" + "\n".join(missing))
    print(f"[截图] ZIP 已保存: {zip_path}")
    if missing:
        print(f"  无截图经销商({len(missing)}家): {', '.join(missing[:10])}")


# ─────────────────────────────────────────────────────────────────────────────
# 主流程
# ─────────────────────────────────────────────────────────────────────────────

def main():
    cfg = TASK_CONFIG
    _autodetect_inputs(cfg)

    if not (cfg.get("dealer_list_xlsx") and
            Path(cfg["dealer_list_xlsx"]).exists()):
        print("\n[错误] 没找到经销商名单 xlsx。")
        print("       请把名单文件（文件名建议含『检核』或『名单』）放到本文件夹后重跑。")
        return

    Path(cfg["output_dir"]).mkdir(parents=True, exist_ok=True)

    print("=" * 60)
    print("  易车奇瑞经销商点检任务")
    print("=" * 60)
    print(f"  名单:    {cfg['dealer_list_xlsx']}")
    print(f"  标准:    {cfg.get('standard_xlsx') or '（未配置，跳过报价对比）'}")
    print(f"  并发数:  {cfg['concurrency']}")
    print("=" * 60)

    dealers = load_dealers_yiche(cfg["dealer_list_xlsx"])
    print(f"[名单] 共 {len(dealers)} 家经销商")
    if not dealers:
        print("[错误] 名单为空，请确认文件格式和易车UID列存在。")
        return

    # 采集
    all_recs = run_scrape(dealers, cfg)

    # 报价对比
    if cfg.get("compare_prices") and cfg.get("standard_xlsx") and \
       Path(cfg["standard_xlsx"]).exists():
        run_compare(cfg["result_xlsx"], cfg["standard_xlsx"], cfg["checked_xlsx"])
    elif cfg.get("compare_prices"):
        print("[核价] standard_xlsx 未配置或不存在，跳过报价对比")

    # 打包截图
    if cfg.get("zip_screenshots") and cfg["features"].get("screenshot"):
        run_zip_screenshots(dealers, cfg)

    print("\n✅ 全部完成")
    out = cfg["checked_xlsx"] if (cfg.get("compare_prices") and
                                   cfg.get("standard_xlsx")) else cfg["result_xlsx"]
    print(f"   主结果: {out}")
    if cfg.get("zip_screenshots"):
        print(f"   截图包: {cfg['screenshots_zip']}")


if __name__ == "__main__":
    main()
